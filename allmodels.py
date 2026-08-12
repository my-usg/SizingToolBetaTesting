import streamlit as st
import sys
import os
import io
from openpyxl import load_workbook

# ── page config ──────────────────────────────────────────────────────────────
st.set_page_config(page_title="General Sizing Tool - All Models", page_icon="⚙️", layout="wide", initial_sidebar_state="collapsed")

st.markdown("<style>[data-testid='stSidebarCollapseButton'],[data-testid='stSidebarCollapsedControl'],[data-testid='collapsedControl']{display:none}</style><h1>General Sizing Tool - All Models</h1>", unsafe_allow_html=True)
st.markdown("Fill in the inputs above and click **Run Sizing**.")

# ── inject all the data + logic from the original script ────────────────────
# We exec the file up to (but not including) the INPUT section so we get all
# the data tables and functions, then drive it with Streamlit widgets.

_script_dir = os.path.dirname(os.path.abspath(__file__))
_tool_path   = os.path.join(_script_dir, "All Models Script.py")

with open(_tool_path, "r") as f:
    _source = f.read()

# Split at line 3587 — the print("ULTIMATE SIZING TOOL") line that starts the I/O section
_lines  = _source.splitlines(keepends=True)
_code   = "".join(_lines[:4068])

_globals = {}
exec(compile(_code, _tool_path, "exec"), _globals)

# Make the tool's functions callable directly in this module
for _k, _v in _globals.items():
    if not _k.startswith("__"):
        globals()[_k] = _v

# Keep a reference so run_tool can update the exec namespace
# (functions defined inside _globals look up variables there, not in globals())


# ── helper ───────────────────────────────────────────────────────────────────
def fmt_pn(pn):
    if isinstance(pn, list):
        return "\n".join(f"• {p}" for p in pn)
    return pn


def build_summary_pdf(inputs, selection, capacity, part_numbers, warnings, adjustments):
    """Build the 'USG Sizing Tool Output' PDF summary and return a BytesIO buffer.

    The layout is compact and auto-scales down until the whole summary fits on a
    single page.

    inputs       : dict of Parameter -> Value
    selection    : list of (label, value) tuples for the chosen regulator
    capacity     : pre-formatted capacity string (or "")
    part_numbers : list or str of HSC part number(s)
    warnings     : list of warning strings
    adjustments  : dict of Adjustment -> Value
    """
    from io import BytesIO
    from datetime import datetime
    from xml.sax.saxutils import escape
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Table, TableStyle)

    ORANGE = colors.HexColor("#e85d26")
    DARK   = colors.HexColor("#111827")
    GREY   = colors.HexColor("#6b7280")
    LINE   = colors.HexColor("#e5e7eb")

    PAGE_W, PAGE_H = letter

    # normalise the section data once
    sel_pairs = [(lbl, val) for lbl, val in selection if val]
    if capacity:
        sel_pairs.append(("Calculated Capacity (CFH)", capacity))
    pns   = part_numbers if isinstance(part_numbers, list) else [part_numbers]
    pns   = [p for p in pns if p]
    warns = [w for w in (warnings or []) if w]

    def _render(scale):
        """Render the whole summary at a given scale; return (buffer, page_count)."""
        base = getSampleStyleSheet()
        title_style = ParagraphStyle("USGTitle", parent=base["Title"], textColor=DARK,
                                     fontSize=18 * scale, leading=20 * scale, spaceAfter=1 * scale, alignment=0)
        sub_style   = ParagraphStyle("USGSub", parent=base["Normal"], textColor=GREY,
                                     fontSize=7.5 * scale, leading=9 * scale, spaceAfter=8 * scale)
        sec_style   = ParagraphStyle("USGSection", parent=base["Heading2"], textColor=ORANGE,
                                     fontSize=10.5 * scale, leading=12 * scale,
                                     spaceBefore=7 * scale, spaceAfter=2 * scale)
        cell_style  = ParagraphStyle("USGCell", parent=base["Normal"],
                                     fontSize=8 * scale, leading=9.5 * scale, textColor=DARK)
        label_style = ParagraphStyle("USGLabel", parent=cell_style, fontName="Helvetica-Bold")

        pad = 2.2 * scale
        margin_x = 0.7 * inch
        margin_y = 0.5 * inch
        avail_w = PAGE_W - 2 * margin_x

        def _grid_style():
            return TableStyle([
                ("VALIGN",        (0, 0), (-1, -1), "TOP"),
                ("LINEBELOW",     (0, 0), (-1, -1), 0.4, LINE),
                ("TOPPADDING",    (0, 0), (-1, -1), pad),
                ("BOTTOMPADDING", (0, 0), (-1, -1), pad),
                ("LEFTPADDING",   (0, 0), (-1, -1), 0),
                ("RIGHTPADDING",  (0, 0), (-1, -1), 0),
            ])

        def kv_table(pairs):
            data = [[Paragraph(escape(str(k)), label_style),
                     Paragraph(escape(str(v)), cell_style)] for k, v in pairs]
            t = Table(data, colWidths=[avail_w * 0.4, avail_w * 0.6])
            t.setStyle(_grid_style())
            return t

        def list_table(items):
            data = [[Paragraph(escape(str(i)), cell_style)] for i in items]
            t = Table(data, colWidths=[avail_w])
            t.setStyle(_grid_style())
            return t

        story = []
        story.append(Paragraph("USG Sizing Tool Output", title_style))
        story.append(Paragraph(
            "United Sales Group \u00b7 Regulator Sizing Platform \u00b7 Generated "
            + datetime.now().strftime("%b %d, %Y  %I:%M %p"), sub_style))

        story.append(Paragraph("Inputs", sec_style))
        story.append(kv_table(list(inputs.items())) if inputs else Paragraph("None", cell_style))

        story.append(Paragraph("Regulator Selection", sec_style))
        story.append(kv_table(sel_pairs) if sel_pairs else Paragraph("No regulator selected.", cell_style))

        story.append(Paragraph("Part Number(s)", sec_style))
        story.append(list_table(pns) if pns else Paragraph("None", cell_style))

        story.append(Paragraph("Warnings", sec_style))
        story.append(list_table(warns) if warns else Paragraph("None", cell_style))

        story.append(Paragraph("Sizing Adjustments", sec_style))
        story.append(kv_table(list(adjustments.items())) if adjustments else Paragraph("None", cell_style))

        buf = BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=letter,
                                topMargin=margin_y, bottomMargin=margin_y,
                                leftMargin=margin_x, rightMargin=margin_x,
                                title="USG Sizing Tool Output")
        doc.build(story)
        buf.seek(0)
        return buf, doc.page

    # Shrink until it fits on one page (or we hit the readability floor).
    buf = None
    for scale in (1.0, 0.95, 0.9, 0.85, 0.8, 0.75, 0.7, 0.65, 0.6):
        buf, pages = _render(scale)
        if pages <= 1:
            break
    return buf

def run_tool(
    inlet, outlet, flow_rate, min_flow, maop,
    pipesize_input, opp_type, irv_input,
    oversizeby, gastypemult, pload, combust_pref
):
    """Run all regulator selection functions and return (result_dict, warnings)."""

    # ── inject globals that the functions read ───────────────────────────────
    # Functions were exec'd into _globals, so that's where they look up names
    _globals["inlet_input"]    = inlet
    _globals["outlet_input"]   = outlet
    _globals["flow_rate"]      = flow_rate
    _globals["min_flow"]       = min_flow
    _globals["maop"]           = maop
    _globals["pipesize_input"] = pipesize_input
    _globals["opp_type"]       = opp_type
    _globals["irv_input"]      = irv_input
    _globals["oversizeby"]     = oversizeby
    _globals["gastypemult"]    = gastypemult
    _globals["pload"]          = pload
    _globals["combust_pref"]   = combust_pref
    _globals["Patm"]           = Patm

    msgs   = []   # warning messages
    result = {}   # what we'll display

    # 143 -------------------------------------------------------------------
    r143, m143, ok143, w143 = run_regulator_selection143(
        inlet, outlet, opp_type)
    if ok143:
        if w143: msgs.append(w143)
        result["match"] = m143
        result["pn"]    = hsc_pnc143(m143)
        return result, msgs

    # 496 -------------------------------------------------------------------
    r496, m496, ok496, w496 = run_regulator_selection496(
        inlet, outlet, opp_type)
    if ok496:
        if w496: msgs.append(w496)
        result["match"]  = m496
        result["pn"]     = hsc_pnc496(m496)
        return result, msgs

    # 243 -------------------------------------------------------------------
    r243, m243, ok243, w243 = run_regulator_selection243(
        inlet, outlet, opp_type)
    if ok243:
        if w243: msgs.append(w243)
        result["match"] = m243
        result["pn"]    = hsc_pnc243(m243)
        return result, msgs

    # 046 -------------------------------------------------------------------
    r046, m046, ok046, w046 = run_regulator_selection046(
        inlet, outlet, opp_type)
    if ok046:
        if w046: msgs.append(w046)
        result["match"] = m046
        result["pn"]    = hsc_pnc046(m046)
        return result, msgs

    # ── new routing: high-eff + no OPP → try 121/122 before 441/461 ────────
    if combust_pref:
        r121, r121vp, r122, m121, ok121, w121 = run_regulator_selection121(
            inlet, outlet, opp_type)
        if ok121:
            if w121: msgs.append(w121)
            result["match"]    = m121
            result["pn"]       = hsc_pnc121(m121)
            result["note121"]  = True
            result["note121_pipe"] = body_size_min121(ip=inlet, reg=m121["reg"])
            return result, msgs

        # 121 didn't work — fall through to 441/461
        m461, ok461, w461 = run_regulator_selection461(
            inlet, outlet, flow_rate, min_flow, opp_type)
        if ok461:
            if w461: msgs.append(w461)
            result["match"] = m461
            result["pn"]    = hsc_pnc461(m461)
            return result, msgs

        result["no_match"] = True
        return result, msgs

    # ── standard routing: 441/461 before 121/122 ────────────────────────────
    m461, ok461, w461 = run_regulator_selection461(
        inlet, outlet, flow_rate, min_flow, opp_type)
    if ok461:
        if w461: msgs.append(w461)
        result["match"] = m461
        result["pn"]    = hsc_pnc461(m461)
        return result, msgs

    # 121/122 ---------------------------------------------------------------
    r121, r121vp, r122, m121, ok121, w121 = run_regulator_selection121(
        inlet, outlet, opp_type)
    if ok121:
        if w121: msgs.append(w121)
        result["match"]    = m121
        result["pn"]       = hsc_pnc121(m121)
        result["note121"]  = True
        result["note121_pipe"] = body_size_min121(ip=inlet, reg=m121["reg"])
        return result, msgs

    result["no_match"] = True
    return result, msgs


# ── SIDEBAR: all 14 inputs ───────────────────────────────────────────────────
def _req_label(key, text):
    """Red label with a ! while a required field is empty; normal once filled."""
    return f":red[{text}] ❗" if not st.session_state.get(key) else text


# ── INPUTS (top of page, always visible) ────────────────────────────────────
st.subheader("📋 Inputs")

_pipe_options = ["N/A", '3/8"', '1/2"', '3/4"', '1"', '1-1/4"', '1-1/2"', '2"', '2-1/2"', '3"']

# Pressures & Flow — values on top row, units/MAOP on the row beneath (aligned grid)
st.markdown("**Pressures & Flow**")
_v1, _v2, _v3, _v4 = st.columns(4)
with _v1:
    inlet_input  = st.number_input(_req_label("k_inlet", "Inlet pressure"),   min_value=0.0, max_value=1000.0, value=0.0, step=0.1, format="%.1f", key="k_inlet")
with _v2:
    outlet_input = st.number_input(_req_label("k_outlet", "Outlet pressure"), min_value=0.0, max_value=1000.0, value=0.0, step=0.1, format="%.1f", key="k_outlet")
with _v3:
    flow_rate    = st.number_input(_req_label("k_flow", "Max gas load / flow rate"), min_value=0, max_value=10000000000, value=0, step=1, format="%d", key="k_flow")
with _v4:
    min_flow_raw = st.number_input("Min gas load / flow rate", min_value=0, max_value=10000000000, value=0, step=1, format="%d", help="Not required: default 0 (enter 0 to use max flow)")

_u1, _u2, _u3, _u4 = st.columns(4)
with _u1:
    inlet_units  = st.selectbox("Inlet pressure units",  ["psi", "bar", "kPa"])
with _u2:
    outlet_units = st.selectbox("Outlet pressure units", ["psi", "in wc", "oz", "bar", "kPa"])
with _u3:
    flowrate_units = st.selectbox("Gas load / flow rate units", ["CFH", "CMH", "BTUH"])
with _u4:
    maop         = st.number_input("Max Allowable Inlet Pressure (psi)", min_value=0, max_value=1000, value=0, step=1, format="%d", help="Not required: default 0.  Regulator sized based on inlet pressure, however program will ensure configuration can handle this max inlet pressure.")
min_flow     = flow_rate if min_flow_raw == 0 else min_flow_raw

# Design Parameters (left)  |  Load Type & Gas (right)
_design, _loadgas = st.columns(2)

with _design:
    st.markdown("**Design Parameters**")
    _sz, _ = st.columns([1, 1])          # keep the pipe-size box compact
    with _sz:
        pipesize_index = st.selectbox("Desired pipe size", range(len(_pipe_options)),
            index=0,
            format_func=lambda i: _pipe_options[i])
    pipesize_input_raw = _pipe_options[pipesize_index]
    pipesize_input = 0 if pipesize_input_raw == "N/A" else pipesize_input_raw

    opp_choice = st.radio("Overpressure protection required?", ["No", "Yes"])
    irv_input  = 0.0
    opp_type   = "None"
    opp_pref   = ""
    if opp_choice == "Yes":
        opp_pref = st.radio("If applicable should the program prioritize sizing with an internal relief valve or default to monitor regulator sizing?", ["IRV (Internal Relief Valve)", "Monitor regulator"])
        if "IRV" in opp_pref:
            irv_input = st.number_input("Internal relief valve should protect downstream pressure to (psi)",
                                        min_value=0.0, max_value=500.0, value=2.0, step=0.1, format="%.1f")
            opp_type = "IRV"
        else:
            opp_type = "Monitor"
    else:
        partial_choice = st.radio("Select regulator with an internal relief valve for partial overpressure protection?", ["No", "Yes"])
        if partial_choice == "Yes":
            opp_type = "Partial"

with _loadgas:
    st.markdown("**Load Type & Gas**")
    higheff   = st.radio("Feeding a generator or high-efficiency boiler?", ["No", "Yes"], help="Program will select a regulator that has capacity for double the load feeding high-efficiency equipment")
    pload     = 0.0
    pload_pct = 0
    if higheff == "Yes":
        pload_pct = st.slider("% of total load feeding high-efficiency appliances", 0, 100, 100, help="Program will select a regulator that has capacity for double the load feeding high-efficiency equipment")
        pload = pload_pct / 100.0
    oversizeby = 1.25 + (0.75 * pload)

    override_oversize = st.radio("Override percentage regulator is oversized by?", ["No", "Yes"], help="Default is set to 25% or 100% for high-efficiency appliances")
    if override_oversize == "Yes":
        oversizeby = st.slider("Oversize regulator by:", 0, 100, 25, help="Recommended to oversize regulator by 20-30%")
        oversizeby = 1 + (oversizeby / 100)
        oversize_percent = (oversizeby - 1) * 100

    combust_pref_choice = st.radio("Prefer combustion regulator (Model 121/122) sizing?", ["No", "Yes"], help="If Yes is selected, the program will attempt to select a Model 121 or 122 regulator before a Model 461 or 441 regulator.")
    combust_pref = combust_pref_choice == "Yes"

    gastype_input = st.selectbox("Gas type", ["Natural Gas", "Propane", "Other"])
    gastypemult   = 1.0
    sg            = 0.6
    if gastype_input == "Propane":
        gastypemult = 0.63
    elif gastype_input == "Other":
        sg = st.number_input("Specific gravity", min_value=0.01, max_value=10.0, value=0.6, step=0.01, format="%.2f")
        gastypemult = min(1.0, (0.6 / sg) ** 0.5)
        st.info("Contact USG for regulator compatibility with gases other than methane or propane.")

    elevation = st.radio("Altitude above 3,000 feet or atmospheric pressure below 13 psi", ["No", "Yes"])
    Patm = 14.4
    if elevation == "Yes":
        Patm  = st.number_input("Atmospheric Pressure (psi)",   min_value=8.80, max_value=14.73, value=14.40,   step=0.01,  format="%.1f")

run_btn = st.button("▶  Run Sizing", type="primary")
st.divider()


# ── MAIN AREA: validation then results ──────────────────────────────────────
# Convert inputs to psi for validation (mirrors the conversion done before run_tool)
def to_psi(val, units):
    if units == "in wc": return val * (1/28)
    if units == "bar":   return val * 14.5
    if units == "oz":    return val / 16
    if units == "kPa":   return val / 6.89476
    return val

inlet_psi  = to_psi(inlet_input, inlet_units)
outlet_psi = to_psi(outlet_input, outlet_units)

# Elevation Reduction Calculation
if Patm < 14.4:
    ratio = (inlet_psi + Patm)/(outlet_psi + Patm)
    if ratio < 1.894:
        elevation_reduction = 100 * (1 - (((outlet_psi+Patm)*((inlet_psi+Patm)-(outlet_psi+Patm)))**0.5) / (((outlet_psi+14.65)*((inlet_psi+14.65)-(outlet_psi+14.65)))**0.5))
    else:
        elevation_reduction = 100 * (1 - (inlet_psi+Patm)/(inlet_psi+14.65))
else:
    elevation_reduction = 0

errors = []
if inlet_psi == 0:
    errors.append("Inlet pressure is required.")
if outlet_psi == 0:
    errors.append("Outlet pressure is required.")
if flow_rate == 0:
    errors.append("Please enter a max gas load / flow rate.")
if inlet_psi > 0 and (inlet_psi > 1000 or inlet_psi < 0.25):
    errors.append("Inlet pressure must be between 7\" wc (0.25 psi / 0.017 bar) and 1,000 psi.")
if outlet_psi > 0 and (outlet_psi < 1.5/28 or outlet_psi > 250):
    errors.append("Outlet pressure must be between 1.5\" wc and 250 psi.")
if inlet_psi > 0 and outlet_psi > 0 and outlet_psi >= inlet_psi:
    errors.append("Outlet pressure must be less than inlet pressure.")
if int(maop) != 0 and maop < inlet_psi:
    errors.append("MAIP must be >= inlet pressure.")
if min_flow > flow_rate:
    errors.append("Minimum flow must be ≤ maximum flow rate.")
if inlet_psi > 0 and outlet_psi > 0 and inlet_psi > 175 and outlet_psi < 3:
    errors.append("Pressure differential too large — consider two pressure cuts.")

if run_btn:
    if errors:
        for e in errors:
            st.error(e)
    else:
        with st.spinner("Sizing regulator…"):
            try:
                # ── unit conversions (mirror the original script) ────────────
                flow_cfh    = flow_rate
                minflow_cfh = min_flow

                # maop defaults to inlet pressure if 0
                maop_psi = inlet_psi if maop == 0 else maop

                if flowrate_units == "CMH":
                    flow_cfh    = flow_rate * 35.3147
                    minflow_cfh = min_flow  * 35.3147
                elif flowrate_units == "BTUH":
                    if gastype_input == "Natural Gas":
                        flow_cfh    = flow_rate / 1000
                        minflow_cfh = min_flow  / 1000
                    elif gastype_input == "Propane":
                        flow_cfh    = flow_rate / 2516
                        minflow_cfh = min_flow  / 2516
                    else:
                        st.error("BTUH conversion is only supported for Natural Gas or Propane. Please enter flow rate in CFH or CMH.")
                        st.stop()              

                result, msgs = run_tool(
                    inlet_psi, outlet_psi, flow_cfh, minflow_cfh, maop_psi,
                    pipesize_input, opp_type, irv_input,
                    oversizeby, gastypemult, pload, combust_pref
                )

                # ── warnings ────────────────────────────────────────────────
                for m in msgs:
                    st.warning(m)

                if result.get("no_match"):
                    st.error("❌  No USG regulators will work for this application.")

                elif "match" in result:
                    match = result["match"]
                    pn    = result["pn"]

                    oversize_percent = (oversizeby - 1) * 100

                    st.success("✅  Regulator selected!")

                    # ── result card ─────────────────────────────────────────
                    st.subheader("Regulator Selection")
                    fields = [
                        ("Model",              match.get("model")),
                        ("Diaphragm Size",     match.get("diap")),
                        ("Body Size",          match.get("body")),
                        ("Orifice Size",       match.get("orifice")),
                        ("Seat",               match.get("seat")),
                        ("Spring",             f"{match.get('color', '')} {match.get('range', '')}".strip()),
                        ("Monitor Spring",     f"{match.get('mon_color','')} {match.get('mon_range','')}".strip() if match.get("mon_color") not in (None, "N/A") else None),
                        ("Monitor Diaphragm",  match.get("mon_diap") if match.get("mon_diap") not in (None, "N/A") else None),
                    ]
                    for label, val in fields:
                        if val:
                            st.markdown(f"**{label}:** {val}")

                    cap = match.get("capacity")
                    if cap and cap != "N/A":
                        try:
                            st.markdown(f"**Calculated Capacity (CFH):** {int(round(float(cap))):,}")
                        except Exception:
                            st.markdown(f"**Calculated Capacity (CFH):** {cap}")

                    st.subheader("HSC Part Number(s)")
                    if isinstance(pn, list):
                        for p in pn:
                            st.code(p)
                    else:
                        st.code(pn)

                    if result.get("note121"):
                        pipe = result.get("note121_pipe", "")
                        if pipe:
                            st.info(f"ℹ️  Model 121 regulators have outlet pipe sizing requirements. This regulator was sized for use with **{pipe}** outlet pipe. For capacities with smaller outlet piping, see regulator brochure.")
                        else:
                            st.info("ℹ️  Model 121 regulators have outlet pipe sizing requirements — see brochure.")

                    # ── sizing adjustments ───────────────────────────────────
                    st.divider()
                    st.subheader("Sizing Adjustments")
                    adj = {"Oversized By": f"{oversize_percent:.0f}%"}
                    if "match" in result and result["match"].get("opp") == "Monitor":
                        adj["Monitor Capacity Reduction"] = "30%"
                    if gastypemult != 1:
                        adj["Gas Type Factor"] = f"{gastypemult:.4f}"
                    if Patm < 14.4:
                        adj["Elevation capacity reduction"] = f"{elevation_reduction:.0f}%"
                    import pandas as pd
                    df_adj = pd.DataFrame(adj.items(), columns=["Adjustment", "Value"])
                    st.dataframe(df_adj, use_container_width=True, hide_index=True)

                    # # ── input summary ─────────────────────────────────────────
                    # st.divider()
                    # st.subheader("Input Summary")
                    summary = {
                        f"Inlet Pressure ({inlet_units})":   inlet_input,
                        f"Outlet Pressure ({outlet_units})": outlet_input,
                        f"Max Flow Rate ({flowrate_units})": f"{flow_rate:,}",
                        f"Min Flow Rate ({flowrate_units})": f"{min_flow:,}",
                        "Max Allowable Inlet Pressure (psi)": f"{int(maop)}",
                        "Requested Pipe Size": _pipe_options[pipesize_index],
                        "Overpressure Protection Required": "Yes" if opp_choice == "Yes" else "No",
                    }
                    if opp_type == "Partial":
                        summary["Select Regulator with IRV"] = "Yes"
                    if opp_choice == "Yes":
                        summary["Protection Type"] = "IRV" if "IRV" in opp_pref else "Monitor"
                        if "IRV" in opp_pref:
                            summary["IRV Protect Downstream Pressure To (psi)"] = f"{irv_input:.1f}"
                    summary["% Load Feeding Generator / High-Eff Boiler"] = f"{pload_pct}%" if higheff == "Yes" else "N/A"
                    summary["Combustion Regulator Preferred"] = "Yes" if combust_pref else "No"
                    summary["Gas Type"] = gastype_input
                    summary["Atmospheric Pressure (psi)"] = f"{Patm:.1f}" if Patm < 14.4 else "14.4"                       
                    # df = pd.DataFrame(summary.items(), columns=["Parameter", "Value"])
                    # st.dataframe(df, use_container_width=True, hide_index=True)

                    # ── PDF summary download ──────────────────────────────────
                    st.divider()
                    st.subheader("Download PDF Summary")
                    try:
                        _cap = match.get("capacity")
                        try:
                            _cap_str = f"{int(round(float(_cap))):,}" if _cap and _cap != "N/A" else ""
                        except Exception:
                            _cap_str = str(_cap) if _cap else ""
                        pdf_buf = build_summary_pdf(
                            inputs=summary,
                            selection=fields,
                            capacity=_cap_str,
                            part_numbers=pn,
                            warnings=msgs,
                            adjustments=adj,
                        )
                        st.download_button(
                            label="⬇️  Download PDF Summary",
                            data=pdf_buf,
                            file_name="USG_Sizing_Tool_Output.pdf",
                            mime="application/pdf",
                        )
                    except Exception as _pex:
                        st.warning(f"Could not generate PDF: {_pex}")

                    # ─────────────────────────────────────────────────────────────
                    # EXCEL / SPREADSHEET DOWNLOAD — DISABLED
                    # The full working code is preserved below, commented out.
                    # To turn it back on, remove the leading "# " from each line
                    # of the block below (and make sure openpyxl is in requirements.txt
                    # and 'Regulator Sizing Examples.xlsx' is in the app directory).
                    # ─────────────────────────────────────────────────────────────
#                     # ── excel download (moved to very bottom; elevation added) ─
#                     st.divider()
#                     st.subheader("Download Summary")
#                     try:
#                         _tmpl_path = os.path.join(_script_dir, "Regulator Sizing Examples.xlsx")
#                         wb = load_workbook(_tmpl_path)
#                         ws = wb.active
#
#                         # ── inputs ──
#                         ws["B2"]  = inlet_units
#                         ws["B3"]  = inlet_input
#                         ws["B4"]  = outlet_units
#                         ws["B5"]  = outlet_input
#                         ws["B6"]  = flowrate_units
#                         ws["B7"]  = flow_rate
#                         ws["B8"]  = min_flow_raw if min_flow_raw > 0 else 0
#                         ws["B9"]  = int(maop)
#                         ws["B10"] = _pipe_options[pipesize_index]
#                         ws["B11"] = "Yes" if opp_choice == "Yes" else "No"
#                         if opp_choice == "Yes":
#                             ws["B12"] = "IRV" if "IRV" in opp_pref else "Monitor"
#                             ws["B13"] = irv_input if "IRV" in opp_pref else ""
#                             ws["B14"] = ""
#                         else:
#                             ws["B12"] = ""
#                             ws["B13"] = ""
#                             ws["B14"] = "Yes" if opp_type == "Partial" else "No"
#                         ws["B15"] = "Yes" if higheff == "Yes" else "No"
#                         ws["B16"] = f"{pload_pct}%" if higheff == "Yes" else ""
#                         ws["B17"] = "Yes" if combust_pref else "No"
#                         ws["B18"] = gastype_input
#                         ws["B19"] = sg if gastype_input == "Other" else ""
#                         # elevation
#                         ws["B20"] = "Yes" if elevation == "Yes" else "No"
#                         ws["B21"] = Patm if elevation == "Yes" else ""
#
#                         # ── outputs (shifted down 2 rows for the elevation rows) ──
#                         ws["B23"] = match.get("model", "")
#                         ws["B24"] = match.get("diap", "") or ""
#                         ws["B25"] = match.get("body", "") or ""
#                         ws["B26"] = match.get("orifice", "") or ""
#                         ws["B27"] = match.get("seat", "") or ""
#                         spring_str = f"{match.get('color','')} {match.get('range','')}".strip()
#                         ws["B28"] = spring_str
#                         mon_color = match.get("mon_color")
#                         mon_range = match.get("mon_range")
#                         if mon_color not in (None, "N/A"):
#                             ws["B29"] = f"{mon_color} {mon_range}".strip()
#                         else:
#                             ws["B29"] = ""
#                         cap = match.get("capacity")
#                         try:
#                             ws["B30"] = int(round(float(cap))) if cap and cap != "N/A" else ""
#                         except Exception:
#                             ws["B30"] = cap or ""
#                         pn_list = pn if isinstance(pn, list) else [pn]
#                         ws["B31"] = pn_list[0] if len(pn_list) > 0 else ""
#                         ws["B32"] = pn_list[1] if len(pn_list) > 1 else ""
#                         warn_list = [m for m in msgs if m]
#                         ws["B33"] = warn_list[0] if len(warn_list) > 0 else ""
#                         ws["B34"] = warn_list[1] if len(warn_list) > 1 else ""
#
#                         buf = io.BytesIO()
#                         wb.save(buf)
#                         buf.seek(0)
#                         st.download_button(
#                             label="⬇️  Download Excel Summary",
#                             data=buf,
#                             file_name="regulator_sizing_summary.xlsx",
#                             mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#                         )
#                     except FileNotFoundError:
#                         st.info("Template file not found — add 'Regulator Sizing Examples.xlsx' to the app directory to enable Excel download.")
#                     except Exception as _ex:
#                         st.warning(f"Could not generate Excel: {_ex}")
#
            except Exception as ex:
                st.error(f"Error during sizing: {ex}")
                import traceback; st.code(traceback.format_exc())

else:
    # placeholder before first run
    st.info("⬆️  Fill in the inputs above and click **Run Sizing**.")